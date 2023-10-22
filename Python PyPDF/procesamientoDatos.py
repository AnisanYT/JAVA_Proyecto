import openpyxl
from openpyxl import load_workbook
import os
import time
import numpy as np

from PyPDF2 import PdfReader

from docx import Document

from PIL import Image
#-----------------------------------

#Aqui le especificamos la ruta de la carpeta a procesar
carpeta = "documentos/"
#Le decimos los archivos en la carpeta
archivos = os.listdir(carpeta)

#-----------------------------------

errores = []

#-----------------------------------
#Asi se llamara el archivo que vamos a crear:
inventario = 'Inventario_Carpetas.xlsx'
rutaInventari = "documentosFin/" + inventario
#Filtramos los archivos excel
archivosExcel = [archivo for archivo in archivos if archivo.endswith('.xlsx')]
procesados = 0
correctamente = 0 
erroneos = 0
if archivosExcel:
        try:
            procesados += 1
            for archivo in archivosExcel:
                rutaReal = "documentos/" + archivo
                workbook = openpyxl.load_workbook(rutaReal)
                activaHoja = workbook.active
                #Nombre, ruta, tamanio
                nombre = os.path.basename(rutaReal)
                ruta = os.path.abspath(rutaReal)
                tamanio = os.path.getsize(rutaReal)
                #metadatos
                properties = workbook.properties
                autor = properties.creator
                fechaCreacion = properties.created
                fechaModificacion = properties.modified
                fechaLastAccess = time.ctime(os.path.getatime(rutaReal))
                fechaLastCreatedTool = properties.lastModifiedBy
                workbook.close()
                tamanioMB = tamanio/1048576
                resultado = "{:.2f} MB".format(tamanioMB)
                #print (f'\n{nombre}, \n{ruta}, \n{resultado}, \n{fechaModificacion}, \n{fechaLastAccess}, \n{fechaLastCreatedTool}')
                #Aqui vamos a verificar primero si esta InventarioDatos, si no esta, entonces la crearemos y posterior a esto
                #vamos a abrir y escribir en el:

                try:
                     workbook = load_workbook(rutaInventari)
                except:
                     workbook = openpyxl.Workbook()

                #Aqui activamos el workbook en el que trabajaremos
                hojaActiva = workbook.active
                #Ahora vamos a escribir en el: 
                hojaActiva.append([""])
                #Tener en cuenta que es una fila, cada posicion es una columna
                hojaActiva.append(["Nombre", nombre])
                hojaActiva.append(["Ruta", ruta])
                hojaActiva.append(["Peso", tamanioMB])
                hojaActiva.append(["Autor", autor])
                hojaActiva.append(["Fecha de creación", fechaCreacion])
                hojaActiva.append(["Fecha de modificacion", fechaModificacion])
                hojaActiva.append(["Fecha de ultimo acceso", fechaLastAccess])
                hojaActiva.append(["Ultima creacion por", fechaLastCreatedTool])
                #gurdamos el archivo
                workbook.save(rutaInventari)
                #cerramos el archivo
                workbook.close()
                correctamente += 1  
        except Exception as a:
            print(f'Sucedio un error: {a}')
            erroneos += 1
            errores.append(a)

            
#Ahora vamos a procesar los archivos docx
#Filtramos los archivos docx
archivosDocx = [archivo for archivo in archivos if archivo.endswith('.docx')]

if archivosDocx:
     procesados += 1
     try:
        for archivo in archivosDocx:
             rutaReal = "documentos/" + archivo
             nombre = archivo
             ruta = os.path.abspath(rutaReal)
             tamanio = os.path.getsize(rutaReal)
             rutaDox = Document(rutaReal)
             properties = rutaDox.core_properties
             autor = properties.author
             fechaCreacion = properties.created
             fechaModificacion = properties.modified

             try:
                  workbook = load_workbook(rutaInventari)
             except: 
                  workbook = openpyxl.Workbook()

             hojaActiva = workbook.active
             hojaActiva.append([""])
             hojaActiva.append(["Nombre", nombre])
             hojaActiva.append(["Ruta", ruta])
             hojaActiva.append(["Tamaño", tamanio])
             hojaActiva.append(["Autor", autor])
             hojaActiva.append(["fecha de cracion", fechaCreacion])
             hojaActiva.append(["Fecha modificacion", fechaModificacion])
             workbook.save(rutaInventari)
             workbook.close()
             correctamente += 1
     except Exception as a:
            print(f'Sucedio un error: {a}')
            erroneos += 1
            errores.append(a)

archivosImagenes = [archivo for archivo in archivos if archivo.endswith(('.jpg', '.jpeg', '.png', '.gif'))]

for archivo in archivosImagenes:
    procesados += 1
    try:
        rutaReal = os.path.join(carpeta, archivo)
        nombre = archivo
        ruta = os.path.abspath(rutaReal)
        tamanio = os.path.getsize(rutaReal)

        # Obtener metadatos de la imagen
        with Image.open(rutaReal) as img:
            ancho, alto = img.size
            formato = img.format

        # Cargar o crear el archivo Excel de inventario
        try:
            workbook = load_workbook(rutaInventari)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # Acceder a la hoja activa
        hojaActiva = workbook.active

        # Agregar los metadatos a la hoja de trabajo
        hojaActiva.append([""])
        hojaActiva.append(["Nombre", nombre])
        hojaActiva.append(["Ruta", ruta])
        hojaActiva.append(["Tamaño", tamanio])
        hojaActiva.append(["Ancho", ancho])
        hojaActiva.append(["Alto", alto])
        hojaActiva.append(["Formato", formato])

        # Guardar el archivo Excel de inventario
        workbook.save(rutaInventari)
        workbook.close()
        correctamente += 1

    except Exception as e:
        print(f'Sucedio un error: {e}')
        erroneos += 1
        errores.append(e)


archivosPDF = [archivo for archivo in archivos if archivo.endswith('.pdf')]
for archivo in archivosPDF:
    procesados += 1
    try:
        rutaReal = os.path.join(carpeta, archivo)
        nombre = archivo
        ruta = os.path.abspath(rutaReal)
        tamanio = os.path.getsize(rutaReal)

        # Obtener metadatos del archivo PDF
        with open(rutaReal, 'rb') as pdf_file:
            pdf_reader = PdfReader(pdf_file)
            num_paginas = len(pdf_reader.pages)

        # Cargar o crear el archivo Excel de inventario
        try:
            workbook = load_workbook(rutaInventari)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # Acceder a la hoja activa
        hojaActiva = workbook.active

        # Agregar los metadatos a la hoja de trabajo
        hojaActiva.append([""])
        hojaActiva.append(["Nombre", nombre])
        hojaActiva.append(["Ruta", ruta])
        hojaActiva.append(["Tamaño", tamanio])
        hojaActiva.append(["Número de Páginas", num_paginas])

        # Guardar el archivo Excel de inventario
        workbook.save(rutaInventari)
        workbook.close()
        correctamente += 1

    except Exception as e:
        print(f'Sucedio un error: {e}')
        erroneos += 1
        errores.append(e)


try:
    workbook = openpyxl.load_workbook(rutaInventari)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

nuevaHoja = workbook.create_sheet(title="Datos")
nuevaHoja['A1'] = 'Correctamente'
nuevaHoja['B1'] = 'Erroneos'
nuevaHoja['C1'] = 'Procesados en total'

nuevaHoja.append([correctamente, erroneos, procesados])

if len(errores) > 0:
    nuevaHoja.append(errores)
else:
    nuevaHoja.append(["Parece que no hubo errores."])


workbook.save(rutaInventari)

workbook.close()
#-----------------------------------
